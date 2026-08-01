import csv
import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]


class ImportTemplateTests(unittest.TestCase):
    def test_listening_workbook_sheets_and_guide(self):
        path = ROOT / 'modules/listening/templates/nghe-mau-day-du.xlsx'
        workbook = load_workbook(path, read_only=True, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                '00_HUONG_DAN',
                '01_BO_TONG_HOP',
                '02_CHI_TU_VUNG',
                '03_CHI_CAU',
                '04_CHI_HOI_THOAI',
                '05_CHI_DOAN_VAN',
                '06_NHOM_NHIEU_BO',
            ],
        )
        guide_values = [cell.value for row in workbook['00_HUONG_DAN'].iter_rows() for cell in row if cell.value]
        guide_text = '\n'.join(map(str, guide_values))
        self.assertIn('deck_id', guide_text)
        self.assertIn('dialogue_turn', guide_text)
        self.assertIn('không có sheet ngữ pháp riêng', guide_text.lower())
        for sheet_name in workbook.sheetnames[1:]:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows())]
            self.assertIn('row_type', headers)
            self.assertIn('deck_id', headers)
            self.assertIn('hanzi', headers)
            self.assertGreater(sheet.max_row, 1)

    def test_flashcard_workbook_sheets_and_guide(self):
        path = ROOT / 'modules/hanzi-stroke/templates/flashcards/the-mau-day-du.xlsx'
        workbook = load_workbook(path, read_only=True, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ['00_HUONG_DAN', '01_MOT_THE', '02_MOT_BO_NHIEU_THE', '03_MOT_NHOM_NHIEU_BO'],
        )
        guide_values = [cell.value for row in workbook['00_HUONG_DAN'].iter_rows() for cell in row if cell.value]
        guide_text = '\n'.join(map(str, guide_values))
        self.assertIn('card', guide_text.lower())
        self.assertIn('deck_id', guide_text)
        for sheet_name in workbook.sheetnames[1:]:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows())]
            self.assertIn('row_type', headers)
            self.assertIn('deck_id', headers)
            self.assertIn('hanzi', headers)
            self.assertGreater(sheet.max_row, 1)

    def test_long_table_csv_txt_and_json_are_utf8_and_structured(self):
        cases = [
            (ROOT / 'modules/listening/templates', 'nghe-mau-day-du'),
            (ROOT / 'modules/hanzi-stroke/templates/flashcards', 'the-mau-day-du'),
        ]
        for folder, stem in cases:
            with self.subTest(stem=stem):
                csv_path = folder / f'{stem}.csv'
                txt_path = folder / f'{stem}.txt'
                json_path = folder / f'{stem}.json'
                with csv_path.open('r', encoding='utf-8-sig', newline='') as stream:
                    rows = list(csv.DictReader(stream))
                self.assertGreater(len(rows), 0)
                self.assertIn('row_type', rows[0])
                self.assertIn('deck_id', rows[0])
                txt_lines = txt_path.read_text(encoding='utf-8-sig').splitlines()
                self.assertGreater(len(txt_lines), 1)
                self.assertIn('\t', txt_lines[0])
                self.assertIn('row_type', txt_lines[0])
                payload = json.loads(json_path.read_text(encoding='utf-8-sig'))
                self.assertIsInstance(payload.get('rows'), list)
                self.assertGreater(len(payload['rows']), 0)

    def test_template_readmes_and_local_vendor_exist(self):
        self.assertTrue((ROOT / 'modules/listening/templates/README.md').is_file())
        self.assertTrue((ROOT / 'modules/hanzi-stroke/templates/flashcards/README.md').is_file())
        vendor = ROOT / 'assets/vendor/jszip.min.js'
        self.assertTrue(vendor.is_file())
        self.assertGreater(vendor.stat().st_size, 10_000)


if __name__ == '__main__':
    unittest.main()
